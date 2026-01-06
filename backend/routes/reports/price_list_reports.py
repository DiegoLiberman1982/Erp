from flask import Blueprint, request, jsonify, send_file
import json
import traceback
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from routes.auth_utils import get_session_with_auth
from utils.http_utils import make_erpnext_request, handle_erpnext_error
from routes.general import get_smart_limit, get_company_abbr, remove_company_abbr

price_list_reports_bp = Blueprint('price_list_reports_bp', __name__)


def _json_ok(data, total=None, filters=None):
    return jsonify({"success": True, "data": data, "total": total or (len(data) if hasattr(data, '__len__') else None), "filters": filters or {}})


def _resolve_company():
    """Try to get the active company from query, headers or cookies.

    Frontend normally sends active company in the X-Active-Company header. Fall
    back to query param `company` or cookie `active_company` if present.
    """
    company = request.args.get('company')
    if company:
        return company
    # Header set by frontend fetchWithAuth
    company = request.headers.get('X-Active-Company')
    if company:
        return company
    # older clients may store in cookie
    company = request.cookies.get('active_company')
    return company


@price_list_reports_bp.route('/api/reports/price-lists/summary', methods=['GET', 'OPTIONS'])
def summary():
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session, headers, user_id, error_response = get_session_with_auth()
    if error_response:
        return error_response

    company = _resolve_company()

    abbr = None
    if company:
        abbr = get_company_abbr(session, headers, company)

    try:
        print('📊 PriceList Summary - Iniciando consulta')

        # Count price lists
        pl_limit = get_smart_limit(company, 'list')
        print(f"🔢 Using smart limits for company='{company}': price_list_limit={pl_limit}")
        params_pl = {'limit_page_length': pl_limit}
        if company:
            params_pl['company'] = company
        print(f"🔗 Fetch Price Lists - Contactando ERPNext...\n📤 Parámetros de query: {params_pl}")
        resp_pl, err_pl = make_erpnext_request(session=session, method='GET', endpoint='/api/resource/Price List', params=params_pl, operation_name='Fetch Price Lists')
        if err_pl:
            return handle_erpnext_error(err_pl, 'Failed fetching price lists')
        pl_data = resp_pl.json()
        price_lists = pl_data.get('data') if isinstance(pl_data, dict) else pl_data
        total_price_lists = len(price_lists) if isinstance(price_lists, list) else 0

        # Fetch selling and buying item prices
        params_sell = {'filters': json.dumps([['selling', '=', 1]]), 'limit_page_length': get_smart_limit(company, 'list'), 'fields': json.dumps(['*'])}
        params_buy = {'filters': json.dumps([['buying', '=', 1]]), 'limit_page_length': get_smart_limit(company, 'list'), 'fields': json.dumps(['*'])}
        if company:
            params_sell['company'] = company
            params_buy['company'] = company

        print(f"🔗 Fetch Selling Item Prices - Contactando ERPNext...\n📤 Parámetros de query: {params_sell}")
        resp_sell, err_sell = make_erpnext_request(session=session, method='GET', endpoint='/api/resource/Item Price', params=params_sell, operation_name='Fetch Selling Item Prices')
        if err_sell:
            return handle_erpnext_error(err_sell, 'Failed fetching selling item prices')
        sell_data = resp_sell.json().get('data', [])
        # Debug: show how many rows returned and how many have item_code
        try:
            sell_with_code = sum(1 for s in sell_data if s and s.get('item_code'))
            print(f"📥 Selling Item Prices: rows={len(sell_data)}, with_item_code={sell_with_code}, sample={sell_data[:3]}")
        except Exception:
            print("� Selling Item Prices: response parse error or unexpected shape")

        print(f"�🔗 Fetch Buying Item Prices - Contactando ERPNext...\n📤 Parámetros de query: {params_buy}")
        resp_buy, err_buy = make_erpnext_request(session=session, method='GET', endpoint='/api/resource/Item Price', params=params_buy, operation_name='Fetch Buying Item Prices')
        if err_buy:
            return handle_erpnext_error(err_buy, 'Failed fetching buying item prices')
        buy_data = resp_buy.json().get('data', [])
        try:
            buy_with_code = sum(1 for b in buy_data if b and b.get('item_code'))
            print(f"📥 Buying Item Prices: rows={len(buy_data)}, with_item_code={buy_with_code}, sample={buy_data[:3]}")
        except Exception:
            print("📥 Buying Item Prices: response parse error or unexpected shape")

        # Unique item counts (one item may have multiple Item Price rows)
        sell_item_codes = set([s.get('item_code') for s in sell_data if s.get('item_code')])
        buy_item_codes = set([b.get('item_code') for b in buy_data if b.get('item_code')])

        # Fetch items (name + item_name) using smart limit and compute totals
        items_limit = get_smart_limit(company, 'list')
        params_items = {'fields': json.dumps(['name','item_name']), 'limit_page_length': items_limit}
        if company:
            params_items['company'] = company
        print(f"🔗 Fetch Items - Contactando ERPNext...\n📤 Parámetros de query: {params_items}")
        resp_items, err_items = make_erpnext_request(
            session=session,
            method='GET',
            endpoint='/api/resource/Item',
            params=params_items,
            operation_name='Fetch Items'
        )
        if err_items:
            return handle_erpnext_error(err_items, 'Failed fetching items')
        items = resp_items.json().get('data', [])

        total_items = len(items)
        # Items without sale price / without purchase price
        items_without_sale = [i for i in items if (i.get('name') not in sell_item_codes)]
        items_without_purchase = [i for i in items if (i.get('name') not in buy_item_codes)]

        # Margins negative: compute by matching sell vs buy rates where both exist
        buy_map = {}
        for b in buy_data:
            code = b.get('item_code')
            if not code: continue
            buy_map.setdefault(code, []).append(b)

        neg_margins = 0
        for s in sell_data:
            code = s.get('item_code')
            if not code or code not in buy_map: continue
            try:
                sale = float(s.get('price_list_rate') or s.get('rate') or s.get('price') or 0)
                buy = float(buy_map[code][0].get('price_list_rate') or buy_map[code][0].get('rate') or buy_map[code][0].get('price') or 0)
                if sale - buy <= 0:
                    neg_margins += 1
            except Exception:
                continue

        # Use unique item counts for selling/buying prices (one item can have multiple price rows)
        total_selling_items = len(sell_item_codes)
        total_buying_items = len(buy_item_codes)

        data = {
            'total_price_lists': total_price_lists,
            'total_items': total_items,
            'total_selling_prices': total_selling_items,
            'total_buying_prices': total_buying_items,
            'items_without_sale_count': len(items_without_sale),
            'items_without_purchase_count': len(items_without_purchase),
            'negative_margins_count': neg_margins,
            # include small samples (name + code) so frontend can show examples if needed
            'items_without_sale_sample': [{'item_code': remove_company_abbr(i.get('name'), abbr) if abbr else i.get('name'), 'item_name': remove_company_abbr(i.get('item_name') or i.get('name'), abbr) if abbr else (i.get('item_name') or i.get('name'))} for i in items_without_sale[:50]],
            'items_without_purchase_sample': [{'item_code': remove_company_abbr(i.get('name'), abbr) if abbr else i.get('name'), 'item_name': remove_company_abbr(i.get('item_name') or i.get('name'), abbr) if abbr else (i.get('item_name') or i.get('name'))} for i in items_without_purchase[:50]],
        }

        # Extra logs requested: counts for debugging
        print(f"📋 Counts -> total_items={total_items}, selling_item_price_count={total_selling_items}, buying_item_price_count={total_buying_items}")
        print(f"📋 Missing -> items_without_sale_count={len(items_without_sale)}, items_without_purchase_count={len(items_without_purchase)}")

        # Log a summary of the payload (not full data) that we will send to frontend
        payload_summary = {
            'total_price_lists': total_price_lists,
            'total_items': total_items,
            'total_selling_prices': total_selling_items,
            'total_buying_prices': total_buying_items,
            'items_without_sale_sample_count': len(data.get('items_without_sale_sample', [])),
            'items_without_purchase_sample_count': len(data.get('items_without_purchase_sample', [])),
        }
        print(f"📤 Payload summary sent to front: {payload_summary}")

        print(f"✅ PriceList Summary - Completado. price_lists={total_price_lists}, items_no_sale={len(items_without_sale)}, neg_margins={neg_margins}")
        return _json_ok(data)

    except Exception as e:
        print('❌ Error en PriceList Summary:', e)
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


# Alias para compatibilidad con llamadas antiguas desde frontend
@price_list_reports_bp.route('/api/reports/summary', methods=['GET', 'OPTIONS'])
def summary_root():
    # Reusar la implementación de summary para mantener un único comportamiento
    return summary()


@price_list_reports_bp.route('/api/reports/price-lists/missing-sale-prices', methods=['GET', 'OPTIONS'])
def missing_sale_prices():
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session, headers, user_id, error_response = get_session_with_auth()
    if error_response:
        return error_response

    price_list = request.args.get('price_list')
    search = request.args.get('search', '').strip().lower()
    company = _resolve_company()

    abbr = None
    if company:
        abbr = get_company_abbr(session, headers, company)

    try:
        print('🔎 Missing Sale Prices - Iniciando')
        # Fetch selling item prices
        params_sell = {'filters': json.dumps([['selling', '=', 1]]), 'limit_page_length': get_smart_limit(company, 'list'), 'fields': json.dumps(['*'])}
        if price_list:
            params_sell['filters'] = json.dumps([['selling', '=', 1], ['price_list', '=', price_list]])
        if company:
            params_sell['company'] = company

        resp_sell, err_sell = make_erpnext_request(session=session, method='GET', endpoint='/api/resource/Item Price', params=params_sell, operation_name='Fetch Selling Item Prices')
        if err_sell:
            return handle_erpnext_error(err_sell, 'Failed fetching selling item prices')
        sell_data = resp_sell.json().get('data', [])
        sell_codes = set([s.get('item_code') for s in sell_data if s.get('item_code')])

        # Fetch items and return those not in sell_codes
        resp_items, err_items = make_erpnext_request(session=session, method='GET', endpoint='/api/resource/Item', params={'fields': json.dumps(['name','item_name','modified','modified_by']), 'limit_page_length': get_smart_limit(company, 'list')}, operation_name='Fetch Items')
        if err_items:
            return handle_erpnext_error(err_items, 'Failed fetching items')
        items = resp_items.json().get('data', [])

        result = []
        for it in items:
            code = it.get('name')
            name = it.get('item_name') or it.get('name')
            if code in sell_codes:
                continue
            if search and search not in (name or '').lower() and search not in (code or '').lower():
                continue
            result.append({'item_code': remove_company_abbr(code, abbr) if abbr else code, 'item_name': remove_company_abbr(name, abbr) if abbr else name, 'modified': it.get('modified'), 'modified_by': it.get('modified_by')})

        print(f"✅ Missing Sale Prices - encontrados: {len(result)}")
        return _json_ok(result, total=len(result), filters={'price_list': price_list, 'search': search})

    except Exception as e:
        print('❌ Error en Missing Sale Prices:', e)
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


@price_list_reports_bp.route('/api/reports/price-lists/missing-purchase-prices', methods=['GET', 'OPTIONS'])
def missing_purchase_prices():
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session, headers, user_id, error_response = get_session_with_auth()
    if error_response:
        return error_response

    price_list = request.args.get('price_list')
    search = request.args.get('search', '').strip().lower()
    company = _resolve_company()

    abbr = None
    if company:
        abbr = get_company_abbr(session, headers, company)

    try:
        print('🔎 Missing Purchase Prices - Iniciando')
        params_buy = {'filters': json.dumps([['buying', '=', 1]]), 'limit_page_length': get_smart_limit(company, 'list'), 'fields': json.dumps(['*'])}
        if price_list:
            params_buy['filters'] = json.dumps([['buying', '=', 1], ['price_list', '=', price_list]])
        if company:
            params_buy['company'] = company

        resp_buy, err_buy = make_erpnext_request(session=session, method='GET', endpoint='/api/resource/Item Price', params=params_buy, operation_name='Fetch Buying Item Prices')
        if err_buy:
            return handle_erpnext_error(err_buy, 'Failed fetching buying item prices')
        buy_data = resp_buy.json().get('data', [])
        buy_codes = set([b.get('item_code') for b in buy_data if b.get('item_code')])

        resp_items, err_items = make_erpnext_request(session=session, method='GET', endpoint='/api/resource/Item', params={'fields': json.dumps(['name','item_name','modified','modified_by']), 'limit_page_length': get_smart_limit(company, 'list')}, operation_name='Fetch Items')
        if err_items:
            return handle_erpnext_error(err_items, 'Failed fetching items')
        items = resp_items.json().get('data', [])

        result = []
        for it in items:
            code = it.get('name')
            name = it.get('item_name') or it.get('name')
            if code in buy_codes:
                continue
            if search and search not in (name or '').lower() and search not in (code or '').lower():
                continue
            result.append({'item_code': remove_company_abbr(code, abbr) if abbr else code, 'item_name': remove_company_abbr(name, abbr) if abbr else name, 'modified': it.get('modified'), 'modified_by': it.get('modified_by')})

        print(f"✅ Missing Purchase Prices - encontrados: {len(result)}")
        return _json_ok(result, total=len(result), filters={'price_list': price_list, 'search': search})

    except Exception as e:
        print('❌ Error en Missing Purchase Prices:', e)
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


@price_list_reports_bp.route('/api/reports/price-lists/price-variance', methods=['GET', 'OPTIONS'])
def price_variance():
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session, headers, user_id, error_response = get_session_with_auth()
    if error_response:
        return error_response

    company = _resolve_company()

    try:
        print('📈 Price Variance - Iniciando')
        params_sell = {'filters': json.dumps([['selling', '=', 1]]), 'limit_page_length': get_smart_limit(company, 'list'), 'fields': json.dumps(['*'])}
        params_buy = {'filters': json.dumps([['buying', '=', 1]]), 'limit_page_length': get_smart_limit(company, 'list'), 'fields': json.dumps(['*'])}
        if company:
            params_sell['company'] = company
            params_buy['company'] = company

        resp_sell, err_sell = make_erpnext_request(session=session, method='GET', endpoint='/api/resource/Item Price', params=params_sell, operation_name='Fetch Selling Item Prices for Variance')
        if err_sell:
            return handle_erpnext_error(err_sell, 'Failed fetching selling item prices')
        sell_data = resp_sell.json().get('data', [])

        resp_buy, err_buy = make_erpnext_request(session=session, method='GET', endpoint='/api/resource/Item Price', params=params_buy, operation_name='Fetch Buying Item Prices for Variance')
        if err_buy:
            return handle_erpnext_error(err_buy, 'Failed fetching buying item prices')
        buy_data = resp_buy.json().get('data', [])

        buy_map = {}
        for b in buy_data:
            code = b.get('item_code')
            if not code: continue
            # pick first by price_list (could improve by choosing specific list)
            buy_map.setdefault(code, []).append(b)

        results = []
        for s in sell_data:
            code = s.get('item_code')
            if not code: continue
            sale_rate = float(s.get('price_list_rate') or s.get('rate') or s.get('price') or 0)
            buy_record = buy_map.get(code, [None])[0]
            buy_rate = float(buy_record.get('price_list_rate') or buy_record.get('rate') or buy_record.get('price') or 0) if buy_record else None
            margin = None
            margin_pct = None
            if buy_rate is not None:
                margin = sale_rate - buy_rate
                if sale_rate:
                    margin_pct = (margin / sale_rate) * 100

            results.append({
                'item_code': code,
                'item_name': s.get('item_name') or s.get('item'),
                'sale': sale_rate,
                'purchase': buy_rate,
                'margin': margin,
                'margin_pct': margin_pct,
                'price_list': s.get('price_list') or s.get('price_list_name')
            })

        print(f"✅ Price Variance - procesados: {len(results)}")
        return _json_ok(results, total=len(results))

    except Exception as e:
        print('❌ Error en Price Variance:', e)
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


@price_list_reports_bp.route('/api/reports/price-lists/recent-updates', methods=['GET', 'OPTIONS'])
def recent_updates():
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session, headers, user_id, error_response = get_session_with_auth()
    if error_response:
        return error_response

    company = _resolve_company()

    abbr = None
    if company:
        abbr = get_company_abbr(session, headers, company)

    try:
        print('🕒 Recent Price Updates - Iniciando')
        params = {'order_by': 'modified desc', 'limit_page_length': get_smart_limit(company, 'list'), 'fields': json.dumps(['*'])}
        if company:
            params['company'] = company

        resp, err = make_erpnext_request(session=session, method='GET', endpoint='/api/resource/Item Price', params=params, operation_name='Fetch Recent Item Price Updates')
        if err:
            return handle_erpnext_error(err, 'Failed fetching recent item price updates')

        data = resp.json().get('data', [])
        result = []
        for d in data:
            typ = 'venta' if d.get('selling') else 'compra' if d.get('buying') else 'venta'
            result.append({
                'item_code': remove_company_abbr(d.get('item_code'), abbr) if abbr else d.get('item_code'),
                'item_name': remove_company_abbr(d.get('item_name') or d.get('item'), abbr) if abbr else (d.get('item_name') or d.get('item')),
                'price_list': d.get('price_list') or d.get('price_list_name'),
                'modified': d.get('modified'),
                'modified_by': d.get('modified_by') or d.get('owner') or user_id,
                'type': typ
            })

        print(f"✅ Recent Price Updates - retornando: {len(result)}")
        return _json_ok(result, total=len(result))

    except Exception as e:
        print('❌ Error en Recent Price Updates:', e)
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


@price_list_reports_bp.route('/api/reports/price-lists/export-sales-xlsx', methods=['GET', 'OPTIONS'])
def export_sales_xlsx():
    """Generate a professionally formatted Excel file for sales price list (customer-facing)."""
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session, headers, user_id, error_response = get_session_with_auth()
    if error_response:
        return error_response

    price_list_name = request.args.get('price_list')
    company = _resolve_company()

    abbr = None
    if company:
        abbr = get_company_abbr(session, headers, company)

    try:
        print(f'📥 Export Sales XLSX - price_list={price_list_name}, company={company}')

        # Fetch company details for branding
        company_data = {}
        if company:
            resp_company, err_company = make_erpnext_request(
                session=session,
                method='GET',
                endpoint=f'/api/resource/Company/{company}',
                operation_name='Fetch Company Details'
            )
            if not err_company:
                company_data = resp_company.json().get('data', {})

        # Fetch selling item prices
        filters = [['selling', '=', 1]]
        if price_list_name:
            filters.append(['price_list', '=', price_list_name])

        params_sell = {
            'filters': json.dumps(filters),
            'limit_page_length': get_smart_limit(company, 'list'),
            'fields': json.dumps(['item_code', 'item_name', 'price_list', 'price_list_rate', 'currency', 'valid_from', 'valid_upto'])
        }
        if company:
            params_sell['company'] = company

        resp_sell, err_sell = make_erpnext_request(
            session=session,
            method='GET',
            endpoint='/api/resource/Item Price',
            params=params_sell,
            operation_name='Fetch Selling Item Prices for Export'
        )
        if err_sell:
            return handle_erpnext_error(err_sell, 'Failed fetching selling item prices')

        sell_data = resp_sell.json().get('data', [])

        # Create Excel workbook with professional styling
        wb = Workbook()
        ws = wb.active
        ws.title = 'Lista de Precios'

        # Define styles
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
        subheader_font = Font(name='Arial', size=11, bold=True)
        subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        data_font = Font(name='Arial', size=10)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Company branding header
        current_row = 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        company_cell = ws[f'A{current_row}']
        company_cell.value = company_data.get('company_name', company or 'Lista de Precios')
        company_cell.font = Font(name='Arial', size=18, bold=True, color='2F75B5')
        company_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 30

        # Company details
        current_row += 1
        if company_data.get('tax_id'):
            ws.merge_cells(f'A{current_row}:E{current_row}')
            tax_cell = ws[f'A{current_row}']
            tax_cell.value = f"CUIT: {company_data.get('tax_id')}"
            tax_cell.font = Font(name='Arial', size=10)
            tax_cell.alignment = Alignment(horizontal='center')
            current_row += 1

        if company_data.get('phone_no') or company_data.get('email'):
            ws.merge_cells(f'A{current_row}:E{current_row}')
            contact_cell = ws[f'A{current_row}']
            contact_parts = []
            if company_data.get('phone_no'):
                contact_parts.append(f"Tel: {company_data.get('phone_no')}")
            if company_data.get('email'):
                contact_parts.append(f"Email: {company_data.get('email')}")
            contact_cell.value = ' | '.join(contact_parts)
            contact_cell.font = Font(name='Arial', size=10)
            contact_cell.alignment = Alignment(horizontal='center')
            current_row += 1

        # Price list title
        current_row += 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = f"LISTA DE PRECIOS: {price_list_name or 'TODAS LAS LISTAS'}"
        title_cell.font = Font(name='Arial', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[current_row].height = 25

        # Date
        current_row += 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        date_cell = ws[f'A{current_row}']
        date_cell.value = f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y')}"
        date_cell.font = Font(name='Arial', size=9, italic=True)
        date_cell.alignment = Alignment(horizontal='center')

        # Empty row
        current_row += 2

        # Column headers
        headers = ['Código', 'Descripción', 'Lista de Precios', 'Precio', 'Moneda']
        for col_idx, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        ws.row_dimensions[current_row].height = 20

        # Data rows
        current_row += 1
        for item in sell_data:
            ws.cell(row=current_row, column=1, value=remove_company_abbr(item.get('item_code', ''), abbr) if abbr else item.get('item_code', ''))
            ws.cell(row=current_row, column=2, value=remove_company_abbr(item.get('item_name', ''), abbr) if abbr else item.get('item_name', ''))
            ws.cell(row=current_row, column=3, value=item.get('price_list', ''))
            
            price_cell = ws.cell(row=current_row, column=4, value=float(item.get('price_list_rate', 0)))
            price_cell.number_format = '#,##0.00'
            
            ws.cell(row=current_row, column=5, value=item.get('currency'))

            # Apply styling to data rows
            for col_idx in range(1, 6):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = data_font
                cell.border = border
                if col_idx in [1, 3, 5]:  # Center align code, price list, currency
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx == 4:  # Right align price
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            # Alternate row colors for readability
            if current_row % 2 == 0:
                fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                for col_idx in range(1, 6):
                    ws.cell(row=current_row, column=col_idx).fill = fill

            current_row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Lista_Precios_{price_list_name or 'Todas'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        print(f"✅ Export Sales XLSX - generado: {len(sell_data)} items")
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print('❌ Error en Export Sales XLSX:', e)
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


@price_list_reports_bp.route('/api/reports/price-lists/export-purchase-xlsx', methods=['GET', 'OPTIONS'])
def export_purchase_xlsx():
    """Generate a standard Excel file for purchase price list (internal use)."""
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session, headers, user_id, error_response = get_session_with_auth()
    if error_response:
        return error_response

    price_list_name = request.args.get('price_list')
    company = _resolve_company()

    abbr = None
    if company:
        abbr = get_company_abbr(session, headers, company)

    try:
        print(f'📥 Export Purchase XLSX - price_list={price_list_name}, company={company}')

        # Fetch buying item prices
        filters = [['buying', '=', 1]]
        if price_list_name:
            filters.append(['price_list', '=', price_list_name])

        params_buy = {
            'filters': json.dumps(filters),
            'limit_page_length': get_smart_limit(company, 'list'),
            'fields': json.dumps(['item_code', 'item_name', 'price_list', 'price_list_rate', 'currency', 'valid_from', 'valid_upto'])
        }
        if company:
            params_buy['company'] = company

        resp_buy, err_buy = make_erpnext_request(
            session=session,
            method='GET',
            endpoint='/api/resource/Item Price',
            params=params_buy,
            operation_name='Fetch Buying Item Prices for Export'
        )
        if err_buy:
            return handle_erpnext_error(err_buy, 'Failed fetching buying item prices')

        buy_data = resp_buy.json().get('data', [])

        # Create Excel workbook with clean styling
        wb = Workbook()
        ws = wb.active
        ws.title = 'Precios de Compra'

        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        data_font = Font(name='Arial', size=10)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Title
        current_row = 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = f"LISTA DE PRECIOS DE COMPRA: {price_list_name or 'TODAS'}"
        title_cell.font = Font(name='Arial', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 25

        # Date
        current_row += 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        date_cell = ws[f'A{current_row}']
        date_cell.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        date_cell.font = Font(name='Arial', size=9)
        date_cell.alignment = Alignment(horizontal='center')

        # Empty row
        current_row += 2

        # Column headers
        headers = ['Código', 'Descripción', 'Lista de Precios', 'Precio', 'Moneda']
        for col_idx, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        ws.row_dimensions[current_row].height = 18

        # Data rows
        current_row += 1
        for item in buy_data:
            ws.cell(row=current_row, column=1, value=remove_company_abbr(item.get('item_code', ''), abbr) if abbr else item.get('item_code', ''))
            ws.cell(row=current_row, column=2, value=remove_company_abbr(item.get('item_name', ''), abbr) if abbr else item.get('item_name', ''))
            ws.cell(row=current_row, column=3, value=item.get('price_list', ''))
            
            price_cell = ws.cell(row=current_row, column=4, value=float(item.get('price_list_rate', 0)))
            price_cell.number_format = '#,##0.00'
            
            ws.cell(row=current_row, column=5, value=item.get('currency'))

            # Apply styling to data rows
            for col_idx in range(1, 6):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = data_font
                cell.border = border
                if col_idx in [1, 3, 5]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx == 4:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            current_row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Precios_Compra_{price_list_name or 'Todas'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        print(f"✅ Export Purchase XLSX - generado: {len(buy_data)} items")
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print('❌ Error en Export Purchase XLSX:', e)
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500

