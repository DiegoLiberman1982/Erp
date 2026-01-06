from flask import Blueprint, request, jsonify, send_file
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from io import BytesIO
from urllib.parse import quote
import calendar
import json

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fpdf import FPDF

from routes.auth_utils import get_session_with_auth
from routes.general import get_active_company, get_smart_limit
from routes.items import get_tax_template_map
from utils.http_utils import make_erpnext_request


iva_reports_bp = Blueprint('iva_reports_bp', __name__)

MONTH_NAMES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}

REPORT_TYPES = {
    'compras': {
        'doctype': 'Purchase Invoice',
        'label': 'Libro IVA Compras',
        'party_doctype': 'Supplier',
        'party_field': 'supplier',
        'party_name_field': 'supplier_name',
        'tax_id_field': 'supplier_tax_id',
        'bill_date_field': 'bill_date'
    },
    'ventas': {
        'doctype': 'Sales Invoice',
        'label': 'Libro IVA Ventas',
        'party_doctype': 'Customer',
        'party_field': 'customer',
        'party_name_field': 'customer_name',
        'tax_id_field': 'customer_tax_id',
        'bill_date_field': 'posting_date'
    }
}

DECIMAL_ZERO = Decimal('0')
TWO_PLACES = Decimal('0.01')
EXCLUDED_LETTER = 'X'

DETAIL_FIELDS = [
    "name", "posting_date", "company", "invoice_type", "voucher_type_code", "voucher_type",
    "punto_de_venta", "invoice_number", "docstatus", "currency", "grand_total", "total",
    "metodo_numeracion_factura_venta", "naming_series", "is_return", "return_against",
    "percepcion_iibb", "percepcion_iva", "net_total", "total_taxes_and_charges",
    "discount_amount", "supplier", "supplier_name", "supplier_tax_id",
    "customer", "customer_name", "customer_tax_id", "title", "posting_time",
    "bill_date", "remarks"
] + [
    "items.item_code", "items.item_name", "items.description", "items.qty", "items.rate",
    "items.amount", "items.base_amount", "items.item_tax_rate", "items.item_tax_template",
    "items.discount_amount"
]


@iva_reports_bp.route('/api/reports/iva', methods=['GET', 'OPTIONS'])
def iva_reports():
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session, headers, user_id, error_response = get_session_with_auth()
    if error_response:
        return error_response

    company = _resolve_company(user_id)
    if not company:
        return jsonify({"success": False, "message": "No se pudo determinar la compania activa"}), 400

    report_type_key = (request.args.get('type') or 'compras').lower()
    if report_type_key not in REPORT_TYPES:
        return jsonify({"success": False, "message": "Tipo de reporte invalido. Use 'compras' o 'ventas'"}), 400

    try:
        month = int(request.args.get('month'))
        year = int(request.args.get('year'))
    except (TypeError, ValueError):
        return jsonify({"success": False, "message": "Debe indicar mes y anio"}), 400

    if month < 1 or month > 12:
        return jsonify({"success": False, "message": "Mes invalido"}), 400
    if year < 2000 or year > 2100:
        return jsonify({"success": False, "message": "Anio invalido"}), 400

    fmt = (request.args.get('format') or 'json').lower()

    try:
        report_payload = _build_report(session, headers, company, report_type_key, month, year)
    except ValueError as ve:
        return jsonify({"success": False, "message": str(ve)}), 400
    except Exception as exc:
        print(f"IVA report error: {exc}")
        return jsonify({"success": False, "message": "No se pudo generar el reporte IVA"}), 500

    if fmt in ('xlsx', 'excel'):
        stream, filename = _build_excel(report_payload)
        return send_file(
            stream,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    if fmt == 'pdf':
        stream, filename = _build_pdf(report_payload)
        return send_file(stream, as_attachment=True, download_name=filename, mimetype='application/pdf')

    return jsonify({"success": True, "data": report_payload})


def _resolve_company(user_id):
    company = request.args.get('company')
    if company:
        return company
    header_company = request.headers.get('X-Active-Company') or request.headers.get('x-active-company')
    if header_company:
        return header_company
    if user_id:
        return get_active_company(user_id)
    return None


def _build_report(session, headers, company, report_type_key, month, year):
    config = REPORT_TYPES[report_type_key]

    start_date = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    end_date = date(year, month, last_day)

    document_names = _fetch_document_names(session, headers, config['doctype'], company, start_date, end_date)
    # Use purchase templates for 'compras' reports and sales templates for 'ventas'
    txn_type = 'purchase' if report_type_key == 'compras' else 'sales'
    tax_map = get_tax_template_map(session, headers, company, transaction_type=txn_type) or {}
    # tax_map is a mapping rate -> template_name
    template_rate_map = {template: Decimal(str(rate)) for rate, template in tax_map.items()}

    party_cache = {}
    address_cache = {}
    rows = []
    totals = {
        'neto': DECIMAL_ZERO,
        'iva': DECIMAL_ZERO,
        'percepcion_iibb': DECIMAL_ZERO,
        'percepcion_iva': DECIMAL_ZERO,
        'total': DECIMAL_ZERO
    }
    skipped_letters = 0

    for doc_name in document_names:
        detail = _fetch_document_detail(session, headers, config['doctype'], doc_name)
        if not detail:
            continue

        letter = _extract_letter(detail)
        if letter == EXCLUDED_LETTER:
            skipped_letters += 1
            continue

        party_info = _get_party_info(
            session,
            headers,
            config,
            detail,
            party_cache,
            address_cache
        )

        breakdown = _build_tax_breakdown(detail, template_rate_map)
        if not breakdown:
            fallback_net = _to_decimal(detail.get('net_total'))
            if fallback_net == DECIMAL_ZERO:
                fallback_net = _to_decimal(detail.get('total'))
            breakdown = {Decimal('0'): fallback_net}

        sorted_breakdown = sorted(breakdown.items(), key=lambda item: float(item[0]))

        percepcion_iibb = _to_decimal(detail.get('percepcion_iibb'))
        percepcion_iva = _to_decimal(detail.get('percepcion_iva'))
        percepcion_consumed = False

        for rate, net_amount in sorted_breakdown:
            iva_amount = (net_amount * rate / Decimal('100')).quantize(TWO_PLACES) if rate is not None else DECIMAL_ZERO

            row_perc_iibb = percepcion_iibb if not percepcion_consumed else DECIMAL_ZERO
            row_perc_iva = percepcion_iva if not percepcion_consumed else DECIMAL_ZERO
            percepcion_consumed = True

            total_row = net_amount + iva_amount + row_perc_iibb + row_perc_iva

            row = {
                "fecha_factura": _format_date(detail.get(config.get('bill_date_field')) or detail.get('posting_date')),
                "fecha_contable": _format_date(detail.get('posting_date')),
                "comprobante": _build_document_label(detail),
                "cuit": detail.get(config['tax_id_field']) or party_info.get('tax_id') or '',
                "razon_social": detail.get(config['party_name_field']) or party_info.get('display_name') or '',
                "condicion_iva": party_info.get('condicion_iva') or 'Sin datos',
                "neto": net_amount.quantize(TWO_PLACES),
                "iva_porcentaje": float(rate) if rate is not None else 0,
                "iva_monto": iva_amount.quantize(TWO_PLACES),
                "provincia": party_info.get('province') or 'Sin datos',
                "percepcion_iibb": row_perc_iibb.quantize(TWO_PLACES),
                "percepcion_iva": row_perc_iva.quantize(TWO_PLACES),
                "total": total_row.quantize(TWO_PLACES),
                "moneda": detail.get('currency')
            }
            rows.append(row)

            totals['neto'] += row['neto']
            totals['iva'] += row['iva_monto']
            totals['percepcion_iibb'] += row['percepcion_iibb']
            totals['percepcion_iva'] += row['percepcion_iva']
            totals['total'] += row['total']

    response_rows = [_serialize_decimal_row(row) for row in rows]
    response_totals = {key: float(value.quantize(TWO_PLACES)) for key, value in totals.items()}

    return {
        "company": company,
        "type": report_type_key,
        "title": f"{config['label']} - {MONTH_NAMES_ES.get(month, month)} {year}",
        "period_label": f"{MONTH_NAMES_ES.get(month, month)} {year}",
        "month": month,
        "year": year,
        "rows": response_rows,
        "totals": response_totals,
        "metadata": {
            "row_count": len(rows),
            "skipped_by_letter_x": skipped_letters,
            "generated_at": datetime.utcnow().isoformat()
        },
        "filters": {
            "company": company,
            "month": month,
            "year": year,
            "type": report_type_key
        }
    }


def _fetch_document_names(session, headers, doctype, company, start_date, end_date):
    filters = [
        ["company", "=", company],
        ["posting_date", ">=", start_date.isoformat()],
        ["posting_date", "<=", end_date.isoformat()],
        ["docstatus", "=", 1]
    ]

    params = {
        "filters": json.dumps(filters),
        "fields": json.dumps(["name"]),
        "order_by": "posting_date asc",
        "limit_page_length": get_smart_limit(company, 'calculate')
    }

    resp, err = make_erpnext_request(
        session=session,
        method='GET',
        endpoint=f"/api/resource/{doctype}",
        params=params,
        operation_name=f"IVA report list for {doctype}"
    )

    if err:
        message = err.get('message') if isinstance(err, dict) else 'No se pudo obtener el listado de comprobantes'
        raise ValueError(message or 'No se pudo obtener el listado de comprobantes')

    if resp.status_code != 200:
        raise ValueError(f"Error al obtener documentos: {resp.status_code}")

    data = resp.json().get('data', [])
    return [row.get('name') for row in data if row.get('name')]


def _fetch_document_detail(session, headers, doctype, doc_name):
    params = {
        "fields": json.dumps(DETAIL_FIELDS)
    }

    resp, err = make_erpnext_request(
        session=session,
        method='GET',
        endpoint=f"/api/resource/{doctype}/{quote(doc_name)}",
        params=params,
        operation_name=f"IVA report detail for {doctype} {doc_name}"
    )

    if err:
        print(f"IVA report detail error for {doc_name}: {err}")
        return None

    if resp.status_code != 200:
        print(f"IVA report detail status {resp.status_code} for {doc_name}")
        return None

    return resp.json().get('data', {})


def _get_party_info(session, headers, config, detail, cache, address_cache):
    party_name = detail.get(config['party_field'])
    if not party_name:
        return {}

    cache_key = (config['party_doctype'], party_name)
    if cache_key in cache:
        return cache[cache_key]

    fields = json.dumps([
        "name", "tax_id", "custom_condicion_iva", "condicion_iva", config['party_name_field'],
        "default_address", "supplier_primary_address", "customer_primary_address",
        "state", "province", "custom_provincia"
    ])

    resp, err = make_erpnext_request(
        session=session,
        method='GET',
        endpoint=f"/api/resource/{config['party_doctype']}/{quote(party_name)}",
        params={"fields": fields},
        operation_name=f"IVA report party {party_name}"
    )

    info = {
        "tax_id": detail.get(config['tax_id_field']),
        "condicion_iva": detail.get('custom_condicion_iva') or detail.get('condicion_iva'),
        "display_name": detail.get(config['party_name_field']) or party_name,
        "province": None
    }

    if not err and resp.status_code == 200:
        data = resp.json().get('data', {})
        info['tax_id'] = data.get('tax_id') or info['tax_id']
        info['condicion_iva'] = data.get('custom_condicion_iva') or data.get('condicion_iva') or info['condicion_iva']
        info['display_name'] = data.get(config['party_name_field']) or info['display_name']
        info['province'] = data.get('state') or data.get('province') or data.get('custom_provincia')

        address_name = data.get('default_address') or data.get('supplier_primary_address') or data.get('customer_primary_address')
        if not info['province'] and address_name:
            info['province'] = _fetch_address_state(session, headers, address_name, address_cache)
    else:
        print(f"IVA report party fetch error for {party_name}: {err}")

    if not info['province']:
        info['province'] = 'Sin datos'

    cache[cache_key] = info
    return info


def _fetch_address_state(session, headers, address_name, cache):
    if address_name in cache:
        return cache[address_name]

    params = {
        "fields": json.dumps(["name", "state", "province", "county", "custom_provincia"])
    }

    resp, err = make_erpnext_request(
        session=session,
        method='GET',
        endpoint=f"/api/resource/Address/{quote(address_name)}",
        params=params,
        operation_name=f"IVA report address {address_name}"
    )

    province = None
    if not err and resp.status_code == 200:
        data = resp.json().get('data', {})
        province = data.get('state') or data.get('province') or data.get('custom_provincia') or data.get('county')
    else:
        print(f"IVA address fetch error for {address_name}: {err}")

    cache[address_name] = province or 'Sin datos'
    return cache[address_name]


def _build_tax_breakdown(detail, template_rate_map):
    breakdown = {}
    items = detail.get('items') or []
    for item in items:
        base_amount = item.get('base_amount')
        if base_amount is None:
            base_amount = item.get('amount')
        net_amount = _to_decimal(base_amount)
        if net_amount == DECIMAL_ZERO:
            continue

        rate = _extract_item_rate(item, template_rate_map)
        rate_key = rate if rate is not None else Decimal('0')
        breakdown[rate_key] = breakdown.get(rate_key, DECIMAL_ZERO) + net_amount

    return breakdown


def _extract_item_rate(item, template_rate_map):
    raw = item.get('item_tax_rate')
    if isinstance(raw, dict) and raw:
        try:
            first_value = next(iter(raw.values()))
            return Decimal(str(first_value))
        except (InvalidOperation, StopIteration):
            pass

    if isinstance(raw, str) and raw.strip():
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict) and parsed:
                first_value = next(iter(parsed.values()))
                return Decimal(str(first_value))
        except (json.JSONDecodeError, InvalidOperation, StopIteration):
            try:
                return Decimal(raw)
            except (InvalidOperation, TypeError):
                pass

    template = item.get('item_tax_template')
    if template and template in template_rate_map:
        return template_rate_map[template]

    return Decimal('0')


def _extract_letter(detail):
    candidates = [
        detail.get('voucher_type_code'),
        detail.get('invoice_type'),
        detail.get('voucher_type'),
        detail.get('metodo_numeracion_factura_venta'),
        detail.get('naming_series')
    ]
    for candidate in candidates:
        letter = _letter_from_value(candidate)
        if letter:
            return letter
    return None


def _letter_from_value(value):
    if not value:
        return None
    upper = str(value).upper()
    if len(upper) == 1 and upper.isalpha():
        return upper

    for token in upper.replace('_', '-').split('-'):
        token = token.strip()
        if len(token) == 1 and token.isalpha():
            return token

    if "FACTURA" in upper:
        parts = upper.split('FACTURA')
        if len(parts) > 1:
            candidate = parts[1].strip().split(' ')[0]
            if len(candidate) == 1 and candidate.isalpha():
                return candidate
    return None


def _build_document_label(detail):
    base = detail.get('voucher_type_code') or detail.get('invoice_type') or detail.get('voucher_type') or 'COMPROBANTE'
    number = _format_document_number(detail.get('punto_de_venta'), detail.get('invoice_number'))
    return f"{base} {number}".strip()


def _format_document_number(punto_de_venta, invoice_number):
    pv = _format_numeric(punto_de_venta, 5)
    num = _format_numeric(invoice_number, 8)
    if pv and num:
        return f"{pv}-{num}"
    return num or pv or ''


def _format_numeric(value, length):
    if value is None:
        return ''
    digits = ''.join(ch for ch in str(value) if ch.isdigit())
    digits = digits or str(value)
    return digits.zfill(length)


def _format_date(value):
    if not value:
        return ''
    for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S.%f'):
        try:
            parsed = datetime.strptime(str(value), fmt)
            return parsed.strftime('%d-%m-%Y')
        except ValueError:
            continue
    return str(value)


def _serialize_decimal_row(row):
    serialized = dict(row)
    for key in ['neto', 'iva_monto', 'percepcion_iibb', 'percepcion_iva', 'total']:
        serialized[key] = float(serialized[key])
    return serialized


def _to_decimal(value):
    if isinstance(value, Decimal):
        return value
    if value is None or value == '':
        return DECIMAL_ZERO
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return DECIMAL_ZERO


def _build_excel(report):
    wb = Workbook()
    ws = wb.active
    ws.title = "Libro IVA"

    accent_dark = "0F172A"
    accent_blue = "1E3A8A"
    header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    total_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    info_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color="D0D7E2"),
        right=Side(style='thin', color="D0D7E2"),
        top=Side(style='thin', color="D0D7E2"),
        bottom=Side(style='thin', color="D0D7E2")
    )

    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = report['title']
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].fill = PatternFill(start_color=accent_dark.replace("#", ""), end_color=accent_dark.replace("#", ""), fill_type="solid")

    ws.merge_cells('A2:G2')
    ws['A2'] = f"Compania: {report['company']}"
    ws['A2'].font = Font(bold=True, color=accent_dark.replace("#", ""))
    ws['A2'].alignment = Alignment(horizontal='left')

    ws.merge_cells('H2:M2')
    ws['H2'] = f"Periodo: {report['period_label']}"
    ws['H2'].font = Font(bold=True, color=accent_dark.replace("#", ""))
    ws['H2'].alignment = Alignment(horizontal='right')

    # KPI row
    summary_metrics = [
        ("Neto", report['totals']['neto']),
        ("IVA", report['totals']['iva']),
        ("Perc. IIBB", report['totals']['percepcion_iibb']),
        ("Perc. IVA", report['totals']['percepcion_iva']),
        ("Total", report['totals']['total']),
    ]
    label_row = 3
    value_row = 4
    col_span = 2
    current_col = 1
    for label, value in summary_metrics:
        end_col = min(current_col + col_span - 1, 13)
        ws.merge_cells(start_row=label_row, start_column=current_col, end_row=label_row, end_column=end_col)
        ws.merge_cells(start_row=value_row, start_column=current_col, end_row=value_row, end_column=end_col)
        label_cell = ws.cell(row=label_row, column=current_col, value=label.upper())
        label_cell.font = Font(size=9, color="64748B", bold=True)
        label_cell.alignment = Alignment(horizontal='center')
        label_cell.fill = info_fill
        value_cell = ws.cell(row=value_row, column=current_col, value=float(value or 0))
        value_cell.font = Font(size=12, bold=True, color=accent_blue.replace("#", ""))
        value_cell.alignment = Alignment(horizontal='center')
        value_cell.fill = info_fill
        value_cell.number_format = '#,##0.00'
        for col in range(current_col, end_col + 1):
            ws.cell(row=label_row, column=col).border = thin_border
            ws.cell(row=label_row, column=col).fill = info_fill
            ws.cell(row=value_row, column=col).border = thin_border
            ws.cell(row=value_row, column=col).fill = info_fill
        current_col += col_span

    ws.append([])

    headers = [
        "Fecha Factura", "Fecha Contable", "Comprobante", "CUIT", "Razon Social",
        "Condicion IVA", "Neto", "% IVA", "IVA", "Provincia", "Perc. IIBB", "Perc. IVA", "Total"
    ]
    ws.append(headers)
    header_row_idx = ws.max_row
    for idx, cell in enumerate(ws[header_row_idx], start=1):
        cell.font = Font(bold=True, color=accent_dark.replace("#", ""))
        cell.fill = header_fill
        align = 'center' if idx <= 6 else 'center'
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = thin_border

    data_start_row = ws.max_row + 1
    for row in report['rows']:
        ws.append([
            row['fecha_factura'],
            row['fecha_contable'],
            row['comprobante'],
            row['cuit'],
            row['razon_social'],
            row['condicion_iva'],
            float(row.get('neto', 0) or 0),
            (float(row.get('iva_porcentaje', 0) or 0) / 100),
            float(row.get('iva_monto', 0) or 0),
            row['provincia'],
            float(row.get('percepcion_iibb', 0) or 0),
            float(row.get('percepcion_iva', 0) or 0),
            float(row.get('total', 0) or 0)
        ])

    totals = report['totals']
    ws.append([
        "", "", "Totales", "", "", "",
        float(totals['neto']),
        "",
        float(totals['iva']),
        "",
        float(totals['percepcion_iibb']),
        float(totals['percepcion_iva']),
        float(totals['total'])
    ])
    totals_row_idx = ws.max_row

    column_widths = [14, 14, 28, 18, 30, 20, 16, 10, 16, 20, 18, 18, 18]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    # numeric formats
    for row in ws.iter_rows(min_row=data_start_row, max_row=totals_row_idx, min_col=1, max_col=13):
        for cell in row:
            if cell.column in (7, 9, 11, 12, 13) and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            if cell.column == 8 and isinstance(cell.value, (int, float)):
                cell.number_format = '0.0%'

    # zebra rows + borders
    for idx, excel_row in enumerate(ws.iter_rows(min_row=data_start_row, max_row=totals_row_idx - 1, min_col=1, max_col=13)):
        fill = zebra_fill if idx % 2 else None
        for cell in excel_row:
            if fill:
                cell.fill = fill
            cell.border = thin_border
            if cell.column <= 6:
                cell.alignment = Alignment(horizontal='left')
            else:
                align = 'right' if cell.column not in (8,) else 'center'
                cell.alignment = Alignment(horizontal=align)

    # Totals row styling
    for cell in ws[totals_row_idx]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right')
    ws.cell(row=totals_row_idx, column=3, value="Totales").alignment = Alignment(horizontal='left', vertical='center')

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{report['title'].replace(' ', '_')}.xlsx"
    return buffer, filename


def _build_pdf(report):
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=16)
    pdf.add_page()

    primary = (15, 23, 42)
    accent = (37, 99, 235)
    zebra_light = (248, 250, 252)
    zebra_dark = (255, 255, 255)

    # Header block
    pdf.set_fill_color(*primary)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 12, _latin(report['title']), ln=1, fill=True)
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 7, _latin(f"Compania: {report['company']}"), ln=1, fill=True)
    pdf.cell(0, 7, _latin(f"Periodo: {report['period_label']}"), ln=1, fill=True)
    pdf.ln(4)

    # KPI chips
    summary_metrics = [
        ("Neto", report['totals']['neto']),
        ("IVA", report['totals']['iva']),
        ("Perc. IIBB", report['totals']['percepcion_iibb']),
        ("Perc. IVA", report['totals']['percepcion_iva']),
        ("Total", report['totals']['total']),
    ]
    usable_width = pdf.w - pdf.l_margin - pdf.r_margin
    gap = 4
    box_count = len(summary_metrics)
    box_width = (usable_width - gap * (box_count - 1)) / box_count
    box_height = 18
    start_y = pdf.get_y()
    x = pdf.l_margin
    for label, value in summary_metrics:
        pdf.set_fill_color(248, 250, 252)
        pdf.set_draw_color(226, 232, 240)
        pdf.rect(x, start_y, box_width, box_height, 'DF')
        pdf.set_xy(x + 3, start_y + 4)
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(box_width - 6, 4, _latin(label.upper()), ln=2)
        pdf.set_font('Helvetica', 'B', 11)
        pdf.set_text_color(15, 23, 42)
        pdf.cell(box_width - 6, 6, _latin(_format_currency(value)), ln=1)
        x += box_width + gap
    pdf.set_y(start_y + box_height + 6)

    columns = [
        ("fecha_factura", "F. Factura", 24),
        ("fecha_contable", "F. Contable", 24),
        ("comprobante", "Comprobante", 46),
        ("cuit", "CUIT", 30),
        ("razon_social", "Razon Social", 45),
        ("condicion_iva", "Cond. IVA", 25),
        ("neto", "Neto", 25),
        ("iva_porcentaje", "% IVA", 16),
        ("iva_monto", "IVA", 25),
        ("provincia", "Provincia", 24),
        ("percepcion_iibb", "Perc. IIBB", 25),
        ("percepcion_iva", "Perc. IVA", 25),
        ("total", "Total", 26)
    ]
    available_width = pdf.w - pdf.l_margin - pdf.r_margin
    base_total = sum(width for _, _, width in columns)
    scale = available_width / base_total if base_total else 1
    columns = [(key, label, round(width * scale, 2)) for key, label, width in columns]
    numeric_keys = {'neto', 'iva_monto', 'percepcion_iibb', 'percepcion_iva', 'total'}

    def render_table_header():
        pdf.set_x(pdf.l_margin)
        pdf.set_fill_color(*accent)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 8)
        for _, label, width in columns:
            pdf.cell(width, 8, _latin(label), border=0, align='C', fill=True)
        pdf.ln(8)
        pdf.set_font('Helvetica', '', 7)
        pdf.set_text_color(15, 23, 42)

    pdf.ln(2)
    render_table_header()

    row_height = 6
    for idx, row in enumerate(report['rows']):
        if pdf.get_y() + row_height > pdf.page_break_trigger:
            pdf.add_page()
            render_table_header()
        fill = zebra_light if idx % 2 else zebra_dark
        pdf.set_fill_color(*fill)
        pdf.set_x(pdf.l_margin)
        for key, _, width in columns:
            value = row.get(key, '')
            if key in numeric_keys:
                text = _format_currency(value)
                align = 'R'
            elif key == 'iva_porcentaje':
                text = f"{float(value or 0):.1f}%"
                align = 'C'
            else:
                text = str(value or '')
                align = 'L'
            pdf.cell(width, row_height, _latin(text), border=0, align=align, fill=True)
        pdf.ln(row_height)
        pdf.set_draw_color(241, 245, 249)
        pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())

    pdf.ln(4)
    totals = report['totals']
    pdf.set_x(pdf.l_margin)
    pdf.set_fill_color(*primary)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 8)
    for idx, (key, _, width) in enumerate(columns):
        if key == 'comprobante':
            text = 'Totales'
            align = 'L'
        elif key == 'neto':
            text = _format_currency(totals['neto'])
            align = 'R'
        elif key == 'iva_monto':
            text = _format_currency(totals['iva'])
            align = 'R'
        elif key == 'percepcion_iibb':
            text = _format_currency(totals['percepcion_iibb'])
            align = 'R'
        elif key == 'percepcion_iva':
            text = _format_currency(totals['percepcion_iva'])
            align = 'R'
        elif key == 'total':
            text = _format_currency(totals['total'])
            align = 'R'
        else:
            text = ''
            align = 'L'
        pdf.cell(width, 8, _latin(text), border=0, align=align, fill=True)
    pdf.ln(10)
    pdf.set_text_color(15, 23, 42)

    pdf_data = pdf.output(dest='S')
    if isinstance(pdf_data, bytearray):
        pdf_bytes = bytes(pdf_data)
    elif isinstance(pdf_data, bytes):
        pdf_bytes = pdf_data
    else:
        pdf_bytes = str(pdf_data).encode('latin-1', 'ignore')

    buffer = BytesIO(pdf_bytes)
    buffer.seek(0)
    filename = f"{report['title'].replace(' ', '_')}.pdf"
    return buffer, filename


def _latin(value):
    if value is None:
        return ''
    if isinstance(value, str):
        return value.encode('latin-1', 'ignore').decode('latin-1')
    return str(value)


def _format_currency(value):
    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)
    formatted = f"{num:,.2f}"
    return formatted.replace(",", "X").replace(".", ",").replace("X", ".")
