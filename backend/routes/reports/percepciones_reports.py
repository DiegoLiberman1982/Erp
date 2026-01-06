"""
Percepciones Reports Module
===========================

Módulo para generar reportes de percepciones sufridas en facturas de compra.
Usa el nuevo modelo de percepciones basado en Purchase Taxes and Charges con custom fields.

Tipos de percepciones:
- INGRESOS_BRUTOS: Agrupadas por provincia (códigos 901-924)
- IVA: Sin desglose por provincia
- GANANCIAS: Sin desglose por provincia
"""

from flask import Blueprint, request, jsonify, send_file
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from io import BytesIO
from urllib.parse import quote
import calendar
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fpdf import FPDF

from routes.auth_utils import get_session_with_auth
from routes.general import get_active_company, get_smart_limit
from utils.http_utils import make_erpnext_request


percepciones_reports_bp = Blueprint('percepciones_reports_bp', __name__)

MONTH_NAMES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}

DECIMAL_ZERO = Decimal('0')
TWO_PLACES = Decimal('0.01')

# Cache para configuración de provincias
_provinces_config = None


def _load_provinces_config():
    """Cargar configuración de provincias desde argentina_perceptions.json"""
    global _provinces_config
    if _provinces_config is not None:
        return _provinces_config
    
    try:
        config_path = os.path.join(
            os.path.dirname(__file__), 
            '..', '..', 'shared', 'argentina_perceptions.json'
        )
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
            _provinces_config = config.get('provinces', {})
            # Filtrar _doc
            _provinces_config = {k: v for k, v in _provinces_config.items() if k != '_doc'}
            return _provinces_config
    except Exception as e:
        print(f"Error loading provinces config: {e}")
        return {}


def _get_province_name(province_code):
    """Obtener nombre de provincia a partir de código AFIP"""
    if not province_code:
        return None
    config = _load_provinces_config()
    return config.get(str(province_code))


@percepciones_reports_bp.route('/api/reports/percepciones', methods=['GET', 'OPTIONS'])
def percepciones_report():
    """
    Endpoint para reporte de percepciones sufridas en compras.
    
    Query params:
    - year: Año (obligatorio)
    - month: Mes (opcional, si no se indica devuelve todo el año)
    - perception_type: Tipo de percepción (opcional: INGRESOS_BRUTOS, IVA, GANANCIAS)
    - province_code: Código de provincia (opcional, solo para IIBB)
    - format: Formato de salida (json, xlsx, pdf)
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session, headers, user_id, error_response = get_session_with_auth()
    if error_response:
        return error_response

    company = _resolve_company(user_id)
    if not company:
        return jsonify({"success": False, "message": "No se pudo determinar la compañía activa"}), 400

    # Parsear parámetros
    try:
        year = int(request.args.get('year'))
    except (TypeError, ValueError):
        return jsonify({"success": False, "message": "Debe indicar año"}), 400

    if year < 2000 or year > 2100:
        return jsonify({"success": False, "message": "Año inválido"}), 400

    month = None
    month_param = request.args.get('month')
    if month_param:
        try:
            month = int(month_param)
            if month < 1 or month > 12:
                return jsonify({"success": False, "message": "Mes inválido"}), 400
        except ValueError:
            return jsonify({"success": False, "message": "Mes inválido"}), 400

    perception_type = request.args.get('perception_type', '').upper() or None
    if perception_type and perception_type not in ('INGRESOS_BRUTOS', 'IVA', 'GANANCIAS'):
        return jsonify({"success": False, "message": "Tipo de percepción inválido"}), 400

    province_code = request.args.get('province_code') or None
    
    fmt = (request.args.get('format') or 'json').lower()

    try:
        report_payload = _build_report(
            session, headers, company, year, month, 
            perception_type, province_code
        )
    except ValueError as ve:
        return jsonify({"success": False, "message": str(ve)}), 400
    except Exception as exc:
        print(f"Percepciones report error: {exc}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": "No se pudo generar el reporte de percepciones"}), 500

    if fmt in ('xlsx', 'excel'):
        stream, filename = _build_excel(report_payload)
        return send_file(
            stream,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    if fmt == 'pdf':
        stream, filename = _build_pdf(report_payload)
        return send_file(stream, as_attachment=True, download_name=filename, mimetype='application/pdf')

    return jsonify({"success": True, "data": report_payload})


def _resolve_company(user_id):
    """Resolver la compañía activa"""
    company = request.args.get('company')
    if company:
        return company
    header_company = request.headers.get('X-Active-Company') or request.headers.get('x-active-company')
    if header_company:
        return header_company
    if user_id:
        return get_active_company(user_id)
    return None


def _build_report(session, headers, company, year, month, perception_type, province_code):
    """Construir el reporte de percepciones"""
    
    # Determinar rango de fechas
    if month:
        start_date = date(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end_date = date(year, month, last_day)
        period_label = f"{MONTH_NAMES_ES.get(month)} {year}"
    else:
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        period_label = f"Año {year}"

    # Obtener todas las Purchase Invoices del período
    invoices = _fetch_invoices_with_perceptions(session, headers, company, start_date, end_date)
    
    rows = []
    totals = {
        'total': DECIMAL_ZERO,
        'by_type': {
            'INGRESOS_BRUTOS': DECIMAL_ZERO,
            'IVA': DECIMAL_ZERO,
            'GANANCIAS': DECIMAL_ZERO
        },
        'by_province': {}
    }
    
    supplier_cache = {}
    
    for invoice in invoices:
        # Obtener taxes de la factura
        taxes = invoice.get('taxes') or []
        
        for tax in taxes:
            # Solo procesar percepciones
            if not tax.get('custom_is_perception'):
                continue
            
            tax_perception_type = tax.get('custom_perception_type', '').upper()
            
            # Filtrar por tipo si se especificó
            if perception_type and tax_perception_type != perception_type:
                continue
            
            tax_province_code = tax.get('custom_province_code') or None
            
            # Filtrar por provincia si se especificó
            if province_code and tax_province_code != province_code:
                continue
            
            # Obtener información del proveedor
            supplier_name = invoice.get('supplier_name') or invoice.get('supplier')
            if invoice.get('supplier') and invoice.get('supplier') not in supplier_cache:
                supplier_info = _fetch_supplier_info(session, headers, invoice.get('supplier'))
                supplier_cache[invoice.get('supplier')] = supplier_info
            else:
                supplier_info = supplier_cache.get(invoice.get('supplier'), {})
            
            # Construir fila
            tax_amount = _to_decimal(tax.get('tax_amount', 0))
            percentage = _to_decimal(tax.get('custom_percentage') or tax.get('rate') or 0)
            province_name = tax.get('custom_province_name') or _get_province_name(tax_province_code)
            
            row = {
                'posting_date': _format_date(invoice.get('posting_date')),
                'document_name': invoice.get('name'),
                'document_label': _build_document_label(invoice),
                'supplier': invoice.get('supplier'),
                'supplier_name': supplier_name or supplier_info.get('supplier_name', ''),
                'supplier_tax_id': invoice.get('tax_id') or supplier_info.get('tax_id', ''),
                'perception_type': tax_perception_type,
                'province_code': tax_province_code,
                'province_name': province_name or '',
                'regimen_code': tax.get('custom_regimen_code') or '',
                'percentage': float(percentage),
                'total_amount': float(tax_amount),
                'description': tax.get('description') or '',
                'account_head': tax.get('account_head') or ''
            }
            rows.append(row)
            
            # Acumular totales
            totals['total'] += tax_amount
            if tax_perception_type in totals['by_type']:
                totals['by_type'][tax_perception_type] += tax_amount
            
            # Acumular por provincia (solo IIBB)
            if tax_perception_type == 'INGRESOS_BRUTOS' and province_name:
                if province_name not in totals['by_province']:
                    totals['by_province'][province_name] = DECIMAL_ZERO
                totals['by_province'][province_name] += tax_amount

    # Ordenar filas por fecha descendente
    rows.sort(key=lambda r: r['posting_date'], reverse=True)
    
    # Serializar totales
    response_totals = {
        'total': float(totals['total'].quantize(TWO_PLACES)),
        'by_type': {k: float(v.quantize(TWO_PLACES)) for k, v in totals['by_type'].items()},
        'by_province': {k: float(v.quantize(TWO_PLACES)) for k, v in totals['by_province'].items()}
    }

    return {
        "company": company,
        "title": f"Percepciones - {period_label}",
        "period_label": period_label,
        "year": year,
        "month": month,
        "rows": rows,
        "totals": response_totals,
        "metadata": {
            "row_count": len(rows),
            "generated_at": datetime.utcnow().isoformat()
        },
        "filters": {
            "company": company,
            "year": year,
            "month": month,
            "perception_type": perception_type,
            "province_code": province_code
        }
    }


def _fetch_invoices_with_perceptions(session, headers, company, start_date, end_date):
    """Obtener facturas de compra con sus taxes"""
    
    filters = [
        ["company", "=", company],
        ["posting_date", ">=", start_date.isoformat()],
        ["posting_date", "<=", end_date.isoformat()],
        ["docstatus", "=", 1]
    ]

    fields = [
        "name", "posting_date", "supplier", "supplier_name", "tax_id",
        "grand_total", "currency"
    ]

    params = {
        "filters": json.dumps(filters),
        "fields": json.dumps(fields),
        "order_by": "posting_date desc",
        "limit_page_length": get_smart_limit(company, 'calculate')
    }

    resp, err = make_erpnext_request(
        session=session,
        method='GET',
        endpoint="/api/resource/Purchase Invoice",
        params=params,
        operation_name="Percepciones report list invoices"
    )

    if err:
        raise ValueError(f"Error obteniendo facturas: {err}")

    if resp.status_code != 200:
        raise ValueError(f"Error al obtener facturas: {resp.status_code}")

    invoices_list = resp.json().get('data', [])
    
    # Ahora obtener los taxes de cada factura
    invoices_with_taxes = []
    for invoice in invoices_list:
        invoice_name = invoice.get('name')
        
        # Obtener taxes de la factura
        taxes_resp, taxes_err = make_erpnext_request(
            session=session,
            method='GET',
            endpoint=f"/api/resource/Purchase Invoice/{quote(invoice_name)}",
            params={"fields": json.dumps(["taxes"])},
            operation_name=f"Fetch taxes for invoice {invoice_name}"
        )
        
        if not taxes_err and taxes_resp.status_code == 200:
            invoice_data = taxes_resp.json().get('data', {})
            taxes = invoice_data.get('taxes', [])
            
            # Solo incluir si tiene percepciones
            has_perceptions = any(
                tax.get('custom_is_perception') 
                for tax in taxes
            )
            
            if has_perceptions:
                invoice['taxes'] = taxes
                invoices_with_taxes.append(invoice)
    
    return invoices_with_taxes


def _fetch_supplier_info(session, headers, supplier_name):
    """Obtener información del proveedor"""
    if not supplier_name:
        return {}
    
    resp, err = make_erpnext_request(
        session=session,
        method='GET',
        endpoint=f"/api/resource/Supplier/{quote(supplier_name)}",
        params={"fields": json.dumps(["name", "supplier_name", "tax_id"])},
        operation_name=f"Fetch supplier info {supplier_name}"
    )
    
    if err or resp.status_code != 200:
        return {}
    
    return resp.json().get('data', {})


def _build_document_label(invoice):
    """
    Construir etiqueta legible del documento.
    El name tiene formato: FE-FAC-A-00004-00000813-00001
    Queremos mostrar: FAC A 00004-00000813
    """
    name = invoice.get('name', '')
    if not name:
        return ''
    
    parts = name.split('-')
    if len(parts) >= 5:
        # parts[0] = FE (factura electrónica) o similar
        # parts[1] = FAC/NDC/NDB (tipo documento)
        # parts[2] = A/B/C (letra)
        # parts[3] = punto de venta
        # parts[4] = número
        tipo = parts[1]
        letra = parts[2]
        pv = parts[3]
        num = parts[4]
        return f"{tipo} {letra} {pv}-{num}"
    
    return name


def _format_date(value):
    """Formatear fecha para display"""
    if not value:
        return ''
    for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S'):
        try:
            parsed = datetime.strptime(str(value), fmt)
            return parsed.strftime('%d-%m-%Y')
        except ValueError:
            continue
    return str(value)


def _to_decimal(value):
    """Convertir a Decimal de forma segura"""
    if isinstance(value, Decimal):
        return value
    if value is None or value == '':
        return DECIMAL_ZERO
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return DECIMAL_ZERO


# ============================================================================
# GENERACIÓN DE EXCEL
# ============================================================================

def _build_excel(report):
    """Generar Excel del reporte de percepciones"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Percepciones"

    accent_dark = "0F172A"
    accent_blue = "1E3A8A"
    header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    total_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    info_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color="D0D7E2"),
        right=Side(style='thin', color="D0D7E2"),
        top=Side(style='thin', color="D0D7E2"),
        bottom=Side(style='thin', color="D0D7E2")
    )

    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = report['title']
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].fill = PatternFill(start_color=accent_dark, end_color=accent_dark, fill_type="solid")

    ws.merge_cells('A2:D2')
    ws['A2'] = f"Compañía: {report['company']}"
    ws['A2'].font = Font(bold=True, color=accent_dark)

    ws.merge_cells('E2:H2')
    ws['E2'] = f"Período: {report['period_label']}"
    ws['E2'].font = Font(bold=True, color=accent_dark)
    ws['E2'].alignment = Alignment(horizontal='right')

    # KPI row - resumen por tipo
    summary_metrics = [
        ("IIBB", report['totals']['by_type'].get('INGRESOS_BRUTOS', 0)),
        ("IVA", report['totals']['by_type'].get('IVA', 0)),
        ("Ganancias", report['totals']['by_type'].get('GANANCIAS', 0)),
        ("TOTAL", report['totals']['total']),
    ]
    
    label_row = 3
    value_row = 4
    col_span = 2
    current_col = 1
    
    for label, value in summary_metrics:
        end_col = current_col + col_span - 1
        ws.merge_cells(start_row=label_row, start_column=current_col, end_row=label_row, end_column=end_col)
        ws.merge_cells(start_row=value_row, start_column=current_col, end_row=value_row, end_column=end_col)
        
        label_cell = ws.cell(row=label_row, column=current_col, value=label.upper())
        label_cell.font = Font(size=9, color="64748B", bold=True)
        label_cell.alignment = Alignment(horizontal='center')
        label_cell.fill = info_fill
        
        value_cell = ws.cell(row=value_row, column=current_col, value=float(value or 0))
        value_cell.font = Font(size=12, bold=True, color=accent_blue)
        value_cell.alignment = Alignment(horizontal='center')
        value_cell.fill = info_fill
        value_cell.number_format = '#,##0.00'
        
        current_col += col_span

    ws.append([])

    # Headers de la tabla
    headers = [
        "Fecha", "Comprobante", "Proveedor", "CUIT", 
        "Tipo", "Provincia", "%", "Importe"
    ]
    ws.append(headers)
    header_row_idx = ws.max_row
    
    for idx, cell in enumerate(ws[header_row_idx], start=1):
        cell.font = Font(bold=True, color=accent_dark)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Datos
    data_start_row = ws.max_row + 1
    type_labels = {
        'INGRESOS_BRUTOS': 'IIBB',
        'IVA': 'IVA',
        'GANANCIAS': 'Ganancias'
    }
    
    for row in report['rows']:
        ws.append([
            row['posting_date'],
            row['document_label'],
            row['supplier_name'],
            row['supplier_tax_id'],
            type_labels.get(row['perception_type'], row['perception_type']),
            row['province_name'] if row['perception_type'] == 'INGRESOS_BRUTOS' else '',
            float(row.get('percentage', 0) or 0),
            float(row.get('total_amount', 0) or 0)
        ])

    # Fila de totales
    ws.append([
        "", "", "", "Totales", "", "",
        "",
        float(report['totals']['total'])
    ])
    totals_row_idx = ws.max_row

    # Anchos de columna
    column_widths = [14, 28, 35, 18, 12, 20, 10, 16]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    # Formatos numéricos y zebra
    for idx, excel_row in enumerate(ws.iter_rows(min_row=data_start_row, max_row=totals_row_idx - 1, min_col=1, max_col=8)):
        fill = zebra_fill if idx % 2 else None
        for cell in excel_row:
            if fill:
                cell.fill = fill
            cell.border = thin_border
            if cell.column == 7:
                cell.number_format = '0.00%'
                cell.alignment = Alignment(horizontal='center')
            elif cell.column == 8:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    # Estilo fila totales
    for cell in ws[totals_row_idx]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = total_fill
        cell.border = thin_border
        if cell.column == 8:
            cell.number_format = '#,##0.00'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{report['title'].replace(' ', '_')}.xlsx"
    return buffer, filename


# ============================================================================
# GENERACIÓN DE PDF
# ============================================================================

def _build_pdf(report):
    """Generar PDF del reporte de percepciones"""
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=16)
    pdf.add_page()

    primary = (15, 23, 42)
    accent = (37, 99, 235)
    zebra_light = (248, 250, 252)
    zebra_dark = (255, 255, 255)

    # Header
    pdf.set_fill_color(*primary)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 12, _latin(report['title']), ln=1, fill=True)
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 7, _latin(f"Compañía: {report['company']}"), ln=1, fill=True)
    pdf.cell(0, 7, _latin(f"Período: {report['period_label']}"), ln=1, fill=True)
    pdf.ln(4)

    # KPI chips
    summary_metrics = [
        ("IIBB", report['totals']['by_type'].get('INGRESOS_BRUTOS', 0)),
        ("IVA", report['totals']['by_type'].get('IVA', 0)),
        ("Ganancias", report['totals']['by_type'].get('GANANCIAS', 0)),
        ("TOTAL", report['totals']['total']),
    ]
    
    usable_width = pdf.w - pdf.l_margin - pdf.r_margin
    gap = 4
    box_count = len(summary_metrics)
    box_width = (usable_width - gap * (box_count - 1)) / box_count
    box_height = 18
    start_y = pdf.get_y()
    x = pdf.l_margin
    
    for label, value in summary_metrics:
        pdf.set_fill_color(248, 250, 252)
        pdf.set_draw_color(226, 232, 240)
        pdf.rect(x, start_y, box_width, box_height, 'DF')
        pdf.set_xy(x + 3, start_y + 4)
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(box_width - 6, 4, _latin(label.upper()), ln=2)
        pdf.set_font('Helvetica', 'B', 11)
        pdf.set_text_color(15, 23, 42)
        pdf.cell(box_width - 6, 6, _latin(_format_currency(value)), ln=1)
        x += box_width + gap
    
    pdf.set_y(start_y + box_height + 6)

    # Columnas
    columns = [
        ("posting_date", "Fecha", 28),
        ("document_label", "Comprobante", 45),
        ("supplier_name", "Proveedor", 60),
        ("supplier_tax_id", "CUIT", 35),
        ("perception_type", "Tipo", 25),
        ("province_name", "Provincia", 35),
        ("percentage", "%", 18),
        ("total_amount", "Importe", 30)
    ]
    
    available_width = pdf.w - pdf.l_margin - pdf.r_margin
    base_total = sum(width for _, _, width in columns)
    scale = available_width / base_total if base_total else 1
    columns = [(key, label, round(width * scale, 2)) for key, label, width in columns]
    
    type_labels = {
        'INGRESOS_BRUTOS': 'IIBB',
        'IVA': 'IVA',
        'GANANCIAS': 'Ganancias'
    }

    def render_table_header():
        pdf.set_x(pdf.l_margin)
        pdf.set_fill_color(*accent)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 8)
        for _, label, width in columns:
            pdf.cell(width, 8, _latin(label), border=0, align='C', fill=True)
        pdf.ln(8)
        pdf.set_font('Helvetica', '', 7)
        pdf.set_text_color(15, 23, 42)

    pdf.ln(2)
    render_table_header()

    row_height = 6
    for idx, row in enumerate(report['rows']):
        if pdf.get_y() + row_height > pdf.page_break_trigger:
            pdf.add_page()
            render_table_header()
        
        fill = zebra_light if idx % 2 else zebra_dark
        pdf.set_fill_color(*fill)
        pdf.set_x(pdf.l_margin)
        
        for key, _, width in columns:
            value = row.get(key, '')
            
            if key == 'total_amount':
                text = _format_currency(value)
                align = 'R'
            elif key == 'percentage':
                text = f"{float(value or 0):.2f}%"
                align = 'C'
            elif key == 'perception_type':
                text = type_labels.get(value, value)
                align = 'C'
            elif key == 'province_name':
                text = value if row.get('perception_type') == 'INGRESOS_BRUTOS' else ''
                align = 'L'
            else:
                text = str(value or '')
                align = 'L'
            
            pdf.cell(width, row_height, _latin(text), border=0, align=align, fill=True)
        
        pdf.ln(row_height)
        pdf.set_draw_color(241, 245, 249)
        pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())

    # Fila de totales
    pdf.ln(4)
    pdf.set_x(pdf.l_margin)
    pdf.set_fill_color(*primary)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 8)
    
    for idx, (key, _, width) in enumerate(columns):
        if key == 'supplier_name':
            text = 'Totales'
            align = 'L'
        elif key == 'total_amount':
            text = _format_currency(report['totals']['total'])
            align = 'R'
        else:
            text = ''
            align = 'L'
        pdf.cell(width, 8, _latin(text), border=0, align=align, fill=True)
    
    pdf.ln(10)

    pdf_data = pdf.output(dest='S')
    if isinstance(pdf_data, bytearray):
        pdf_bytes = bytes(pdf_data)
    elif isinstance(pdf_data, bytes):
        pdf_bytes = pdf_data
    else:
        pdf_bytes = str(pdf_data).encode('latin-1', 'ignore')

    buffer = BytesIO(pdf_bytes)
    buffer.seek(0)
    filename = f"{report['title'].replace(' ', '_')}.pdf"
    return buffer, filename


def _latin(value):
    """Convertir a latin-1 para PDF"""
    if value is None:
        return ''
    if isinstance(value, str):
        return value.encode('latin-1', 'ignore').decode('latin-1')
    return str(value)


def _format_currency(value):
    """Formatear valor como moneda"""
    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)
    formatted = f"{num:,.2f}"
    return formatted.replace(",", "X").replace(".", ",").replace("X", ".")
